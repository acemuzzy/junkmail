__author__ = 'Murray'

import imaplib
import email
import re
import openpyxl
from openpyxl.cell import get_column_letter
import os
import yaml

WORKSHEET_NAME = "CURRENT_CONTENT"
BLACKLIST_NAME = "BLACK_LIST"

COLUMNS = [
    "SENDER", "ID"
]


class ExcelHandler:
    """
    Class to handle Excel interactions.
    """

    def __init__(self, filename):
        """
        Constructor.

        :param filename: Which Excel file to use.
        """
        self.filename = filename
        self.blacklist = []
        self.blacklist_row = 2
        self.wb = None

    def _get_wb(self):
        if not self.wb:
            self.wb = openpyxl.load_workbook(self.filename)
        return self.wb

    def exists(self):
        """
        Does the Excel file exist?
        """
        print("Checking existence")
        return os.path.exists(self.filename)

    def create_file(self):
        """
        Create the Excel file, with sensible column headers.

        Note: should only be called if it doesn't exist already!
        """
        print "Creating file"

        self.wb = openpyxl.Workbook()
        ws = self.wb.get_active_sheet()
        ws.title = WORKSHEET_NAME

        for col, colName in enumerate(COLUMNS):
            cell = ws.cell(row=1, column=(col+1))
            cell.value = colName
            cell.font = openpyxl.styles.Font(bold=True)

        ws.column_dimensions[get_column_letter(2)].width = 50
        ws.freeze_panes = ws['B2']

        ws = self.wb.create_sheet(BLACKLIST_NAME)
        ws.column_dimensions[get_column_letter(1)].width = 50
        ws.freeze_panes = ws['B2']

        print " Done!"
        exit(0)

    def read_blacklist(self):
        """
        Read the members of the BLACKLIST_NAME tab.
        """
        print("Reading blacklist")

        wb = self._get_wb()
        ws = wb.get_sheet_by_name(BLACKLIST_NAME)
        ws.column_dimensions[get_column_letter(1)].width = 50

        while True:
            name = ws.cell(row=self.blacklist_row, column=1).value

            if name:
                if name not in self.blacklist:
                    self.blacklist.append(name)
            else:
                break

            self.blacklist_row += 1

        print("Blacklist is {}".format(self.blacklist))

    def update_blacklist(self):
        """
        Update the members of the BLACKLIST_NAME tab with any names marked with
        and 'x' on the main message tab.
        """
        print("Updating blacklist")
        wb = self._get_wb()
        ws = wb.get_sheet_by_name(WORKSHEET_NAME)
        bs_ws = wb.get_sheet_by_name(BLACKLIST_NAME)
        read_row = 2

        while True:
            x = ws.cell(row=read_row, column=1).value
            name = ws.cell(row=read_row, column=2).value

            if name:
                if (name not in self.blacklist) and (x == "x"):
                    self.blacklist.append(name)
                    bs_ws.cell(row=self.blacklist_row, column=1).value = name
                    print("Putting {} at {}".format(name, self.blacklist_row))
                    self.blacklist_row += 1
            else:
                break

            read_row += 1

        print("Blacklist is now {}".format(self.blacklist))

    def do_work(self, sender_list):
        """
        Do actual Excel work, parsing messages etc.
        """
        print("Outputting data")
        wb = self._get_wb()
        ws = wb.get_sheet_by_name(WORKSHEET_NAME)

        output_row = 2
        for sender in sender_list:
            ws.cell(row=output_row, column=2).value = sender

            if sender in self.blacklist:
                ws.cell(row=output_row, column=1).value = "x"
                print("delete {}".format(sender))
            else:
                ws.cell(row=output_row, column=1).value = ""
                print("leave {}".format(sender))

            output_col = 3
            for uid in sender_list[sender]["ids"]:
                if sender in self.blacklist:
                    print("delete id {}".format(uid))
                ws.cell(row=output_row, column=output_col).value = uid
                ws.column_dimensions[get_column_letter(output_col)].width = 6
                output_col += 1

            output_row += 1

        # clear any other rows
        while True:
            if ws.cell(row=output_row, column=1).value or \
               ws.cell(row=output_row, column=2).value or \
               ws.cell(row=output_row, column=3).value:
                ws.cell(row=output_row, column=1).value = ""
                ws.cell(row=output_row, column=2).value = ""

                output_col = 3
                while True:
                    if not ws.cell(row=output_row, column=output_col).value:
                        break
                    ws.cell(row=output_row, column=output_col).value = None
                    output_col += 1

            else:
                break

            output_row += 1

        ws.column_dimensions[get_column_letter(2)].width = 50

        wb.save(self.filename)


class ImapHandler:
    """
    Class for handling IMAP interactions.
    """
    def __init__(self):
        """
        Constructor.
        """
        self._mail = None

    def connect(self, server, username, password):
        """
        Create connection to the IMAP server.

        :param server: The server to connect to.
        :param username: The username to use for the connection.
        :param password: The password to use for the connection.
        """
        try:
            self._mail = imaplib.IMAP4_SSL(server)
            self._mail.login(username, password)

        except imaplib.IMAP4.error as exc:
            print("IMAP error: {}".format(exc))
            exit(0)

    def read_inbox(self):
        """
        Parase the inbox and do actual message handling.
        """
        self._mail.select("inbox")

        result, data = self._mail.search(None, "ALL")
        ids = data[0]
        id_list = ids.split()
        print("{} emails".format(len(id_list)))
        ids = id_list[-50:]

        sender_list = {}

        for uid in ids:
            result, data = self._mail.fetch(uid, "(RFC822)")
            for response_part in data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_string(response_part[1])
                    sender = msg['from'].split()[-1]
                    address = re.sub(r'[<>]', '', sender)

                    print("{},{}".format(uid, address))

                    if address in sender_list:
                        sender_list[address]["ids"].append(uid)
                    else:
                        sender_list[address] = {}
                        sender_list[address]["ids"] = [uid]

        for sender in sender_list:
            print("{}: {}".format(sender, sender_list[sender]["ids"]))

        return sender_list


class ConfigHandler:
    """
    Config handler class.  Deals with parsing YAML to get IMAP / Excel details.
    """

    def __init__(self, filename="junk.cfg"):
        """
        Constructor.

        :param filename: filename to read
        """
        print("Loading configuration")

        with open(filename, 'r') as stream:
            try:
                doc = yaml.load(stream)

                imap_details = doc["imap_server"]
                self.server = imap_details["server"]
                self.username = imap_details["username"]
                self.password = imap_details["password"]

                excel_details = doc["excel"]
                self.filename = excel_details["filename"]

            except yaml.YAMLError as exc:
                print("YAML error: {}".format(exc))
                exit(0)
            except KeyError as exc:
                print("Missing key: {}".format(exc))
                exit(0)

        print("Config loaded")

# Main function
if __name__ == "__main__":
    config = ConfigHandler()

    imap_handler = ImapHandler()
    imap_handler.connect(config.server, config.username, config.password)

    writer = ExcelHandler(config.filename)

    if not writer.exists():
        writer.create_file()
    else:
        writer.read_blacklist()
        writer.update_blacklist()

    senders = imap_handler.read_inbox()
    writer.do_work(senders)